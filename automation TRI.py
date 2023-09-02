import tkinter as tk
from tkinter import font
from tkinter import filedialog
from threading import Thread
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from openpyxl import Workbook, load_workbook, workbook
import time
import os
import pandas as pd
from PIL import Image, ImageTk
from tabulate import tabulate
from prettytable import PrettyTable


class AutomationApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Automation Create by Rama")

        custom_font = font.Font(family="Arial Black", size=18, weight="bold")
        custom_font_button = font.Font(family="Arial", size=9, weight="bold")
        self.header_label = tk.Label(
            root,
            text="Automation Data",
            bg="#f12666",
            fg="#fff",
            anchor="center",
            height="2",
            font=custom_font,
        )

        icon = Image.open("excel.ico")
        self.icon = ImageTk.PhotoImage(icon)

        self.input_button = tk.Button(
            root,
            text="Load Excel Data",
            command=self.load_excel_data,
            bg="#f12666",
            fg="#fff",
            height="2",
            font=custom_font_button,
        )
        self.start_button = tk.Button(
            root,
            text="Start Automation",
            command=self.start_automation,
            bg="#f12666",
            fg="#fff",
            height="2",
            font=custom_font_button,
        )
        self.delete_button = tk.Button(
            root,
            text="Delete Excel Data",
            command=self.delete_excel_data,
            bg="#f12666",
            fg="#fff",
            height="2",
            font=custom_font_button,
        )

        self.excel_button = tk.Button(
            root,
            text=" Export to Excel ",
            image=self.icon,
            compound="left",
            command=self.export_to_excel,
            state=tk.DISABLED,
            bg="#f12666",
            fg="#fff",
            font=custom_font_button,
        )
        self.file_label = tk.Label(root, text="No Excel file loaded")
        self.message_label = tk.Label(
            root,
            text="",
            fg="red",
            anchor="center",
        )  # Label untuk pesan error
        self.log_text = tk.Text(root, height=10, width=130)

        # Create the log_text widget with adjusted width and height
        log_text_width = 100  # Set the desired width for the log text
        log_text_height = 20  # Set the desired height for the log text
        self.log_text = tk.Text(
            root,
            width=log_text_width,
            height=log_text_height,
            wrap="word",
        )

        # Menggunakan tata letak grid untuk tombol-tombol dan elemen lainnya
        self.header_label.grid(row=0, column=0, columnspan=2, sticky="ew")
        self.input_button.grid(row=1, column=0, padx=5, pady=5, sticky="ew")
        self.delete_button.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        self.start_button.grid(row=2, column=0, padx=5, pady=5, sticky="ew")
        self.excel_button.grid(row=2, column=1, padx=5, pady=5, sticky="ew")
        self.file_label.grid(row=3, column=0, columnspan=2, sticky="w")
        self.message_label.grid(row=4, column=0, columnspan=2, sticky="w")
        self.log_text.grid(row=5, column=0, columnspan=2, padx=5, pady=5, sticky="nsew")

        # Konfigurasi tata letak grid agar responsif saat diperbesar atau diperkecil
        for i in range(7):
            self.root.grid_rowconfigure(i, weight=2)
        for i in range(2):
            self.root.grid_columnconfigure(i, weight=1)

        self.excel_path = None
        self.excel_data = None

    def load_excel_data(self):
        self.excel_data = None
        self.excel_path = filedialog.askopenfilename(
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if self.excel_path:
            try:
                self.excel_data = load_workbook(filename=self.excel_path)["Sheet1"]
                self.file_label.config(
                    text="Excel File Name: " + os.path.basename(self.excel_path)
                )
                self.message_label.config(text="")  # Hapus pesan error jika ada
                self.log_text.delete("1.0", tk.END)
                self.log_text.insert(tk.END, "Excel data loaded.\n")
                self.excel_button.config(state=tk.NORMAL)
                self.start_button.config(state=tk.NORMAL)
            except Exception as e:
                self.log_text.insert(
                    tk.END, "Error loading Excel data: {}\n".format(str(e))
                )
                self.excel_data = None
                self.file_label.config(text="No Excel file loaded")
                self.message_label.config(
                    text="Error loading Excel data",
                    anchor="center",
                )  # Tampilkan pesan error
                self.log_text.delete("1.0", tk.END)
                self.excel_button.config(state=tk.DISABLED)
                self.start_button.config(state=tk.DISABLED)

    def export_to_excel(self):
        if self.excel_data is None:
            self.log_text.insert(tk.END, "No Excel data to export.\n")
            return

        output_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")]
        )

        if output_path:
            try:
                # Mengambil data dari tabel log di log_text
                log_text_contents = self.log_text.get("1.0", tk.END)
                table_data = []
                for line in log_text_contents.split("\n"):
                    if line.startswith("|"):
                        table_data.append(
                            line.split("|")[1:-1]
                        )  # Memisahkan data kolom

                # Menyimpan data dalam DataFrame pandas
                df = pd.DataFrame(
                    table_data,
                    columns=["NOMOR", "DATA1", "DATA2", "DATA3", "DATA4", "STATUS"],
                )

                # Menyimpan DataFrame sebagai file Excel
                df.to_excel(output_path, index=False)

                self.log_text.insert(
                    tk.END, "Data exported to Excel: {}\n".format(output_path)
                )
            except Exception as e:
                self.log_text.insert(
                    tk.END, "Error while exporting to Excel: {}\n".format(str(e))
                )

    def start_automation(self):
        if self.excel_data is None:
            self.log_text.insert(tk.END, "Please load Excel data first.\n")
        else:
            automation_thread = Thread(target=self.run_automation)
            automation_thread.start()

    def delete_excel_data(self):
        self.excel_data = None
        self.file_label.config(text="No File")
        self.log_text.delete("1.0", tk.END)  # Clear the log text
        self.log_text.insert(tk.END, "Excel data has been deleted.\n")
        self.start_button.config(
            state=tk.DISABLED
        )  # Nonaktifkan tombol Start Automation

    def run_automation(self):
        wb = load_workbook(filename=self.excel_path)
        sheetRange = wb["Sheet1"]

        driver = webdriver.Chrome()
        driver.get("https://")
        driver.maximize_window()
        driver.implicitly_wait(10)

        self.log_text.insert(tk.END, "Starting Automation...\n")

        i = 2

        table_data = []  # Menyimpan data untuk tabel
        table_headers = [
            "NOMOR",
            "DATA1",
            "DATA2",
            "DATA3",
            "DATA4",
            "STATUS",
        ]  # Header tabel

        while i <= sheetRange.max_row:
            nomor = sheetRange["A" + str(i)].value
            data1 = sheetRange["B" + str(i)].value
            data2 = sheetRange["C" + str(i)].value
            data3 = sheetRange["D" + str(i)].value

            if nomor is not None:
                try:
                    WebDriverWait(driver, 10).until(
                        EC.visibility_of_element_located(
                            (
                                By.XPATH,
                                "<--ISI XPATH-->",
                            )
                        )
                    )
                    driver.find_element(
                        By.XPATH,
                        "<--ISI XPATH-->",
                    ).click()

                    WebDriverWait(driver, 10).until(
                        EC.visibility_of_element_located(
                            (
                                By.XPATH,
                                "<--ISI XPATH-->",
                            )
                        )
                    )
                    driver.find_element(
                        By.XPATH,
                        "<--ISI XPATH-->",
                    ).send_keys(nomor)
                    time.sleep(0.1)
                    driver.find_element(
                        By.XPATH,
                        "<--ISI XPATH-->",
                    ).send_keys(data1)
                    time.sleep(0.1)
                    driver.find_element(
                        By.XPATH,
                        "<--ISI XPATH-->",
                    ).send_keys(data2)
                    time.sleep(0.1)
                    driver.find_element(By.ID, "mat-radio-3").click()
                    time.sleep(0.1)
                    driver.find_element(
                        By.XPATH,
                        "<--ISI XPATH-->",
                    ).send_keys(data3)
                    time.sleep(0.1)
                    driver.find_element(
                        By.XPATH,
                        "<--ISI XPATH-->",
                    ).click()
                    time.sleep(0.1)
                    driver.find_element(
                        By.XPATH,
                        "<--ISI XPATH-->",
                    ).click()
                    time.sleep(0.1)
                    element = driver.find_element(By.CLASS_NAME, "detailstatus")
                    element = element.text
                    element1 = driver.find_element(By.CLASS_NAME, "detailpadding")
                    element1 = element1.text
                    driver.find_element(By.CLASS_NAME, "mat-button-wrapper").click()
                    time.sleep(0.1)

                except TimeoutException:
                    self.log_text.insert(
                        tk.END, "Form didn't appear for row {}\n".format(i - 1)
                    )
                row_data = (
                    nomor,
                    data1,
                    data2,
                    data3,
                    element,
                    element1,
                )  # Menyimpan data baris sebagai tuple
                table_data.append(
                    row_data
                )  # Menambahkan tuple ke dalam list table_data
                self.log_text.insert(tk.END, "'{}',".format(i - 1))
                self.log_text.insert(tk.END, "'{}',".format(nomor))
                self.log_text.insert(tk.END, "'{}',".format(data1))
                self.log_text.insert(tk.END, "'{}',".format(data2))
                self.log_text.insert(tk.END, "'{}',".format(data3))
                self.log_text.insert(tk.END, "'{}',".format(element))
                self.log_text.insert(tk.END, "{}.\n".format(element1))
                self.log_text.see(tk.END)
                # self.log_text.config(state=tk.DISABLED)  # Kembali ke mode hanya-read
                time.sleep(0.1)

            i += 1

        driver.quit()

        log_table = PrettyTable()
        log_table.field_names = table_headers
        log_table.add_rows(table_data)

        # Menampilkan tabel log di log_text
        formatted_table = tabulate(table_data, headers=table_headers, tablefmt="grid")
        self.log_text.insert(tk.END, "Complete Automation.\n")
        self.log_text.insert(tk.END, "\n")
        self.log_text.insert(tk.END, formatted_table + "\n")


if __name__ == "__main__":
    root = tk.Tk()
    app = AutomationApp(root)
    root.mainloop()
